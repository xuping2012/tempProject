# coding:utf-8

from openpyxl import load_workbook
import config
from common.HandleLogging import log

class CasesData(object):
    '''用来保存用例数据的类'''
    pass


class HandleExcel(object):
    '''读取excel工具类。'''
    #行和列都是从1开始数起
    
    def __init__(self,filename,sheetname=None):
        '''实例化文件属性，初始化操作文件对象'''
        self.filename=filename
        self.sheetname=sheetname
#         创建打开文件对象,在构造函数中初始化，在执行测试用例，结果写入时，有多个sheet结果写入，只会在执行最后一个sheet有写入结果
#         self.wb=load_workbook(self.filename)
#         默认sheetname为None，传第一个sheet
#         self.ws=self.wb[self.sheetname] if self.sheetname is not None else self.wb.active

    def open(self):
        '''打开工作簿及表单对象'''
        try:
    #         创建工作簿对象
            self.wb=load_workbook(self.filename)
    #         默认sheetname为None，传第一个sheet
            self.ws=self.wb[self.sheetname] if self.sheetname is not None else self.wb.active
        except Exception as e:
            log.error("{}文件不存在".format(self.filename))
            raise e 
        
        
    def close(self):
        '''关闭工作簿对象'''
        self.wb.close()


    def get_all_cases(self):
        '''获取excel所有行的测试用例'''
        """读取数据"""
        self.open()
        # 获取最大的行
        max_row = self.ws.max_row
        # 读取所有的数据，放到一个列表中
        list_data = []
        for i in range(1, max_row + 1):
            data1 = self.ws.cell(row=i, column=1).value
            data2 = self.ws.cell(row=i, column=2).value
            data3 = self.ws.cell(row=i, column=3).value
            data4 = self.ws.cell(row=i, column=4).value
            list_data.append([data1, data2, data3, data4])
        
        self.close()
        # 创建一个字典，用来存储所有的用来数据
        cases = []
        # 获取第一行数据中的表头
        title = list_data[0]
        for data in list_data[1:]:
            # 遍历第一行之外的其他行数据，聚合打包成字典
            case = dict(zip(title,data))
            cases.append(case)
        
        return cases
    
    
    def get_one_case(self,row):
        '''获取指定行的测试用例数据，row为case_id顺位'''
        # 打开工作簿对象
        self.open()
        # 获取最大的列和行，最小行
        max_col=self.ws.max_column
        max_row=self.ws.max_row
        min_row=self.ws.min_row
        #获取所有行
        rows=list(self.ws.rows)
        #关闭工作簿对象
        self.close()
        
        title=[]#表头
        for row in rows[0]:
            title.append(row.value)
        
        list_data=[]
#         先判断行在范围内
        if isinstance(row,int) and min_row<=row<=max_row:
            for j in range(1,max_col+1):
                data1 = self.ws.cell(row=row+1, column=j).value #提取每个单元格的数据，重新组成一个列表
                print(data1)
                list_data.append(data1)
            case=dict(zip(title,list_data))
            return case
        else:
            log.error( "你输入的行号:{}不正确!".format(str(row)))
    
    
    def get_title_col_index(self,kword):
        '''获取表头关键字参数，返回索引，便于结合行列准确写入结果
                    用于python37版本后，有序的list
        '''
        #打开
        self.open()
#         获取所有行
        rows=list(self.ws.rows)
        #关闭
        self.close()
        
        title=[]#表头
        for row in rows[0]:
            title.append(row.value)
            
        if kword in title:
            return title.index(kword)+1
        else:
            log.error("title中不存在该表头：{}".format(kword))
            
        #虽然可以少一些代码，但是运算速度会低一些
#         for i in range(len(rows[0])):
#             #遍历表头，返回列的索引
#             if rows[0][i].value == kword:
#                 return i+1
    
    
    def get_data_row_index(self,data):
        '''获取执行用例所在行，用于回写结果于哪一行
                    用于python37版本后，有序的list
        '''
#         获取所有测试用例数据
        datas=self.get_datas()
#         再查出元素所在列表的索引，
        if data in datas:
            return datas.index(data)+2 #获取行的索引是从0开始，第0行是title，在读取测试用例数据的行本身就是从1开始，所以+2
        else:
            log.error('datas没有该行数据:{}'.format(data))
        
        
    def get_datas(self):
        '''获取excel所有行的测试用例'''
        """读取数据"""
        self.open()
        # 获取表单所有行
        rows=list(self.ws.rows)
        # 读取工作簿表单所有行数据之后，就可以关闭
        self.close()
        #接收表头的空列表
        title=[]
        for t in rows[0]: #0,第一行就是title
            title.append(t.value)
        
        #接收所有测试用例数据的空列表
        cases=[]
        for row in rows[1:]: #截取list表头后面的元素，遍历
            data=[]
            for r in row:
                #遍历每一行的单元格的值加入一个空列表
                data.append(r.value)
                #zip打包函数
            case=dict(zip(title,data)) #一行测试用例数据，数据格式是dict的items类型，可以转成dict
            #每一行测试用例数据追加成一个用例集合
            cases.append(case)
        return cases
       
                
    def get_datas_obj(self):
        '''获取excel所有行的测试用例'''
        """读取数据"""
        self.open()
        # 获取表单所有行
        rows=list(self.ws.rows)
        # 关闭工作簿对象
        self.close()
        #接收表头的空列表
        title=[]
        for t in rows[0]: #0,第一行就是title
            title.append(t.value)
        #定义空列表用例集合
        cases=[]
        for row in rows[1:]: #截取list表头后面的元素，遍历
            data=[]
            for r in row:
                #遍历每一行的单元格的值加入一个空列表
                data.append(r.value)
            #一行测试用例数据，元组元素组成的列表数据,zip返回的是迭代器,list接收,长度取最短的列表打包成元组元素的list
            case=list(zip(title,data)) 
            #创建对象用来保存一行的用例数据
            case_obj=CasesData()
            #遍历列表中该行的元素，是个元组，k，v元组拆包
            for k,v in case:
                setattr(case_obj, k, v) #设置对象的属性及属性值
            #追加一行测试对象到用例集合列表
            cases.append(case_obj)
            
        log.info("类对象存储测试数据")
        return cases
    
    
    def write_data(self,data,kword,result_status):
        '''
                     执行测试用例，根据用例所在行，写入结果列，并保存
        data:读取excel遍历的测试数据case：
        kword:读取excel表头的索引
        message:写入excel的内容
        return:
        '''
        row=self.get_data_row_index(data)
        column=self.get_title_col_index(kword)
        
        if isinstance(row, int) and isinstance(column, int):
            self.write_file(row,column,result_status)
        else:
            log.info("获取的行:{}或者列:{}不是数字".format(row,column))
            
            
    def write_file(self,row,column,result_status):
        '''执行用例结果写入excel，并保存'''
        if isinstance(row, int) and isinstance(column, int):
            try:
                self.open()
                self.ws.cell(row=row,column=column,value=result_status)#self.get_col_kw(kword)
            except Exception as e:
                log.info("写入异常")
                raise e
            else:
                self.wb.save(self.filename)
#                 log.info("写入成功")
            finally:
                self.close()
        else:
            log.info("输入的行:{}或者列:{}不是数字".format(row,column))

    
if __name__ == '__main__':
    excel = HandleExcel(config.datas_path+"test.xlsx",sheetname='register')
    datas=excel.get_datas()
    
    data=datas[0]
    row=excel.get_data_row_index(data)
    print(row,data)
    excel.write_data(data, "result", "通过")
#     for r in excel.get_datas():
#         print(r.__dict__)
#     excel.write_data(3, "result", "通过")
#     print(excel.get_title_index("result"))